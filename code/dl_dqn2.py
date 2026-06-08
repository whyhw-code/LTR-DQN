import pandas as pd
import torch as T
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
import numpy as np
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

class Environment():
    def __init__(self, data, temp_data, start_date, end_date):
        self.data = data
        self.temp_data = temp_data

        self.data = self.data.fillna(0)
        self.temp_data = self.temp_data.fillna(0)
        self.data['qid_date'] = pd.to_numeric(self.data['qid_date'].str.replace('-', ''))
        self.data['trade_date'] = pd.to_numeric(self.data['trade_date'].str.replace('-', ''))
        # 提取需要归一化的列
        features_to_normalize = self.data.drop(columns=['qid_date', 'trade_date'])
        # 计算最小值和最大值
        min_vals = features_to_normalize.min()
        max_vals = features_to_normalize.max()
        # 进行 [-1, 1] 标准化处理
        normalized_features = 2 * (features_to_normalize - min_vals) / (max_vals - min_vals) - 1
        # 将归一化后的数据合并回原 DataFrame
        self.data = self.data[['qid_date', 'trade_date']].join(normalized_features)

        self.start_date = start_date
        self.end_date = end_date
        self.qid_date = 20220930

        self.buy_fee_rate = 0.0003
        self.sell_fee_rate = 0.0003 + 0.001

        self.init = 500_000_000
        self.fund = 500_000_000

        self.total_profit = 0
        self.day_profit = 0
        self.select_num = 0
        self.day_real_return = 0
        self.select_positive_count = 0

    def reset(self):
        self.qid_date = self.start_date

        self.init = 500_000_000  #初始资金
        self.fund = 500_000_000 #剩余资金

        self.total_profit = 0  #总收益
        self.day_profit = 0  #单日收益
        self.select_num = 0
        self.day_real_return = 0
        self.select_positive_count = 0

        observation = self.data[self.data['trade_date'] == self.qid_date].values.flatten().tolist()
        observation.append(self.day_real_return/0.10)
        observation.append((self.select_num-2)/2)
        observation.append((self.select_positive_count-2)/2)
        return (observation)

    def step(self, action_):
        #
        selected_df = self.temp_data[self.temp_data['qid_date'] == self.qid_date]
        # 基于 prediction 由高到低排序
        selected_df = selected_df.sort_values(by='prediction', ascending=False)
        # 选择前 action 个研报数据
        if action_ > len(selected_df):
            action_ = len(selected_df)
        if action_ > 0:
            selected_df = selected_df.head(action_)
            # self.day_real_return = selected_df['real_return'].mean()-(self.buy_fee_rate+self.sell_fee_rate)
            self.select_positive_count = (selected_df['real_return'] > 0).sum()

            capital_per_stock = self.fund / action_
            self.day_profit = 0
            for _, row in selected_df.iterrows():
                # 计算能买的股数（向下取整）
                shou_num = int(capital_per_stock / (100 * row['pclose']))
                buyfei = shou_num * 100 * row['pclose'] * self.buy_fee_rate
                shares_bought = int((capital_per_stock - self.buy_fee_rate) / (100 * row['pclose'])) * 100
                yu = capital_per_stock - buyfei - shares_bought * row['pclose']
                if shares_bought == 0:
                    print(f"Warning: Not enough capital to buy any shares of stock {row['stock_code']} on {self.qid_date}.")
                    yu = capital_per_stock
                # 计算卖出后的资金
                sell_value = shares_bought * row['close'] - shares_bought * row['close'] * (self.sell_fee_rate) + yu
                # 累加当日总收益
                self.day_profit += (sell_value - capital_per_stock)
            self.day_real_return = self.day_profit / self.fund

        else:
            self.day_real_return = 0.0
            self.select_positive_count = 0
            self.day_profit = 0.0
        self.select_num = action_
        # self.day_profit = self.fund * self.day_real_return
        self.fund = self.fund + self.day_profit
        self.total_profit += self.day_profit

        self.qid_date = self.data[self.data['qid_date'] == self.qid_date]['trade_date'].iloc[0]
        observation_ = self.data[self.data['trade_date'] == self.qid_date].values.flatten().tolist()
        observation_.append(self.day_real_return/0.10)
        observation_.append((self.select_num-2)/2)
        observation_.append((self.select_positive_count-2)/2)

        reward = self.day_real_return
        if self.qid_date == self.end_date:
            done = True
        else:
            done = False

        return (observation_, reward, done, action_, self.select_positive_count)


class DeepQNetwork(nn.Module):
    def __init__(self, lr, input_dims, fc1_dims, fc2_dims, n_actions):
        super(DeepQNetwork, self).__init__()
        self.lr = lr
        self.input_dims = input_dims
        self.fc1_dims = fc1_dims
        self.fc2_dims = fc2_dims
        self.n_actions = n_actions

        self.fc1 = nn.Linear(*self.input_dims, self.fc1_dims)
        self.fc2 = nn.Linear(self.fc1_dims, self.fc2_dims)
        self.fc3 = nn.Linear(self.fc2_dims, self.n_actions)
        # self.optimizer = torch.optim.SGD(self.parameters(), lr=lr, momentum=0.9)
        # self.optimizer = optim.RMSprop(self.parameters(), lr=lr, alpha=0.9)#设置优化器
        self.optimizer = optim.Adam(self.parameters(), lr=lr, betas=(0.9, 0.999), eps=1e-08)
        self.loss = nn.MSELoss()

        self.device = T.device('cuda:0' if T.cuda.is_available() else 'cpu')
        self.to(self.device)

    def forward(self, state):
        state.to(self.device)
        x1 = F.tanh(self.fc1(state.to(T.float32)))
        x2 = F.tanh(self.fc2(x1))
        # x = F.tanh(self.fc2(x))
        actions = self.fc3(x2)

        return actions


class Agent():
    # gamma的折扣率它必须介于0和1之间。越大，折扣越小。这意味着学习，agent 更关心长期奖励。另一方面，gamma越小，折扣越大。这意味着我们的 agent 更关心短期奖励（最近的奶酪）。
    # epsilon探索率ϵ。即策略是以1−ϵ的概率选择当前最大价值的动作，以ϵ的概率随机选择新动作。
    def __init__(self, gamma, epsilon, lr, input_dims, batch_size, n_actions=5,
                 max_mem_size=100, eps_end=0.03, eps_dec=0.0002, fc1_dims=256,fc2_dims=128):
        self.replace_target_iter = 8
        self.learn_step_counter = 0
        self.gamma = gamma
        self.epsilon = epsilon
        self.eps_min = eps_end
        self.eps_dec = eps_dec
        self.lr = lr
        self.n_actions = n_actions
        self.action_space = [i for i in range(n_actions)]
        self.mem_size = max_mem_size
        self.batch_size = batch_size
        self.mem_cntr = 0

        self.Q_eval = DeepQNetwork(self.lr, input_dims=input_dims, n_actions=self.n_actions,
                                   fc1_dims=fc1_dims, fc2_dims=fc2_dims)

        self.Q_target = DeepQNetwork(self.lr, input_dims=input_dims, n_actions=self.n_actions,
                                   fc1_dims=fc1_dims, fc2_dims=fc2_dims)

        self.state_memory = np.zeros((self.mem_size, *input_dims), dtype=np.float32)
        self.new_state_memory = np.zeros((self.mem_size, *input_dims), dtype=np.float32)

        self.action_memory = np.zeros(self.mem_size, dtype=np.int32)
        self.reward_memory = np.zeros(self.mem_size, dtype=np.float32)
        self.terminal_memory = np.zeros(self.mem_size, dtype=bool)

    def save_model(self, model_path):
        T.save(self.Q_eval, model_path)

    def load_model(self, model_path):
        self.Q_eval = T.load(model_path)

    # 存储记忆
    def store_transition(self, state, action, reward, state_, done):

        index = self.mem_cntr % self.mem_size
        self.state_memory[index] = state
        self.new_state_memory[index] = state_
        self.reward_memory[index] = reward
        self.action_memory[index] = action
        self.terminal_memory[index] = done

        self.mem_cntr += 1
        print("store_transition index:", index)

    # observation就是状态state
    def choose_action(self, observation):
        r = np.random.random()

        if r > self.epsilon:
            print('not random action')
            # 随机0-1，即1-epsilon的概率执行以下操作,最大价值操作
            state = T.tensor(observation).to(self.Q_eval.device)
            # 放到神经网络模型里面得到action的Q值vector
            actions = self.Q_eval.forward(state)
            action = T.argmax(actions).item()
        else:
            # epsilon概率执行随机动作
            action = np.random.choice(self.action_space)
            print("random action:", action)
        return action

    def _replace_target_params(self):
        # 复制网络参数
        print(self.Q_eval)
        print(self.Q_target)
        self.Q_target.load_state_dict(self.Q_eval.state_dict())

    # 从记忆中抽取batch进行学习
    def learn(self):
        # memory counter小于一个batch大小的时候直接return
        if self.mem_cntr < self.batch_size:
            print("learn:watching")
            return 0.0000

        if self.learn_step_counter % self.replace_target_iter == 0:
            self._replace_target_params()
            print('\ntarget params replaced\n')

        # 初始化梯度0
        self.Q_eval.optimizer.zero_grad()

        # 得到memory大小，不超过mem_size
        max_mem = min(self.mem_cntr, self.mem_size)

        # 随机生成一个batch的memory index，不可重复抽取
        batch = np.random.choice(max_mem, self.batch_size, replace=False)

        # int序列array，0~batch_size
        batch_index = np.arange(self.batch_size, dtype=np.int32)

        # 从state memory中抽取一个batch
        state_batch = T.tensor(self.state_memory[batch]).to(self.Q_eval.device)
        new_state_batch = T.tensor(self.new_state_memory[batch]).to(self.Q_eval.device)
        reward_batch = T.tensor(self.reward_memory[batch]).to(self.Q_eval.device)
        terminal_batch = T.tensor(self.terminal_memory[batch]).to(self.Q_eval.device)  # 存储是否结束的bool型变量

        # action_batch = T.tensor(self.action_memory[batch]).to(self.Q_eval.device)
        action_batch = self.action_memory[batch]

        # 第batch_index行，取action_batch列,对state_batch中的每一组输入，输出action对应的Q值,batchsize行，1列的Q值
        q_eval = self.Q_eval.forward(state_batch)[batch_index, action_batch]
        # q_next = self.Q_eval.forward(new_state_batch)
        q_next = self.Q_target(new_state_batch)
        q_next[terminal_batch] = 0.0  # 如果是最终状态，则将q值置为0
        q_target = reward_batch + self.gamma * T.max(q_next, dim=1)[0]
        loss = self.Q_eval.loss(q_target, q_eval).to(self.Q_eval.device)
        loss.backward()
        self.Q_eval.optimizer.step()

        self.epsilon = self.epsilon - self.eps_dec if self.epsilon > self.eps_min \
            else self.eps_min
        self.learn_step_counter += 1
        return loss.item()

class T4ExcelWriter:
    """
    将 T4 结果写入 ../result/batch123/基准+模型结果对比.xlsx

    当前逻辑：
    - 不修改原始 results_df
    - 不修改回测过程
    - 只在写入 Excel 时，将资金结果同比缩小 5 倍
    - 保留 Excel 原有格式，只改对应单元格的值
    """

    def __init__(self, excel_path=None, scale=5):
        if excel_path is None:
            self.excel_path = (
                Path(__file__).resolve().parents[1]
                / "result"
                / "batch123"
                / "基准+模型结果对比.xlsx"
            )
        else:
            self.excel_path = Path(excel_path)

        self.scale = scale

        self.sheet_map = {
            "0060": "主板",
            "3068": "创业板",
        }

    def _to_int_date(self, value):
        if pd.isna(value):
            return None

        if hasattr(value, "strftime"):
            return int(value.strftime("%Y%m%d"))

        if isinstance(value, str):
            value = value.strip()
            if value == "":
                return None

            # 关键：用 pandas 正确解析 2021/12/7 这种非补零日期
            parsed_date = pd.to_datetime(value)
            return int(parsed_date.strftime("%Y%m%d"))

        return int(float(value))

    def _get_col_index(self, ws, header_name):
        for cell in ws[1]:
            if cell.value == header_name:
                return cell.column

        raise ValueError(
            f"Cannot find column '{header_name}' in sheet '{ws.title}'."
        )

    def write(
            self,
            results_df,
            dapan_code,
            target_col,
            value_col=None,
            date_col="qid_date",
            count_col=None,
            count_target_col="number of stocks",
            initial_date=20211206,
            start_date=20211207,
            initial_value=1_000_000,
    ):
        if dapan_code not in self.sheet_map:
            raise ValueError(f"Unsupported dapan_code: {dapan_code}")

        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")

        if value_col is None:
            if "total_profit" in results_df.columns:
                value_col = "total_profit"
            elif "funds" in results_df.columns:
                value_col = "funds"
            else:
                raise ValueError(
                    "Cannot infer value_col. Expected 'total_profit' or 'funds'."
                )

        if date_col not in results_df.columns:
            raise ValueError(f"results_df does not contain date column: {date_col}")

        if value_col not in results_df.columns:
            raise ValueError(f"results_df does not contain value column: {value_col}")

        write_df = results_df.copy()

        write_df["_qid_date_key"] = write_df[date_col].map(self._to_int_date)
        write_df = write_df.dropna(subset=["_qid_date_key"])
        write_df["_qid_date_key"] = write_df["_qid_date_key"].astype(int)

        # 核心：只在写入 Excel 时同比缩小 5 倍
        values = write_df[value_col].astype(float) / self.scale
        write_df = write_df[write_df["_qid_date_key"] >= start_date]
        value_map = dict(zip(write_df["_qid_date_key"], values))

        count_map = None
        if count_col is not None:
            if count_col not in write_df.columns:
                raise ValueError(
                    f"results_df does not contain count column: {count_col}"
                )

            count_map = dict(
                zip(
                    write_df["_qid_date_key"],
                    write_df[count_col].astype(int)
                )
            )

        wb = load_workbook(self.excel_path)

        sheet_name = self.sheet_map[dapan_code]
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {self.excel_path}")

        ws = wb[sheet_name]

        excel_date_col = self._get_col_index(ws, "qid_date")
        target_col_idx = self._get_col_index(ws, target_col)

        count_target_col_idx = None
        if count_map is not None:
            count_target_col_idx = self._get_col_index(ws, count_target_col)

        updated = 0
        updated_count = 0

        for row in range(2, ws.max_row + 1):
            qid_date = self._to_int_date(
                ws.cell(row=row, column=excel_date_col).value
            )

            if qid_date == initial_date:
                ws.cell(row=row, column=target_col_idx).value = initial_value
                updated += 1

            elif qid_date in value_map:
                ws.cell(row=row, column=target_col_idx).value = float(value_map[qid_date])
                updated += 1

            if count_map is not None:
                if qid_date == initial_date:
                    ws.cell(row=row, column=count_target_col_idx).value = 0
                    updated_count += 1

                elif qid_date in count_map:
                    ws.cell(row=row, column=count_target_col_idx).value = int(count_map[qid_date])
                    updated_count += 1
                updated_count += 1

        wb.save(self.excel_path)

        print(
            f"[T4ExcelWriter] Updated {updated} rows: "
            f"sheet={sheet_name}, column={target_col}, scale=1/{self.scale}"
        )

        if count_map is not None:
            print(
                f"[T4ExcelWriter] Updated {updated_count} rows: "
                f"sheet={sheet_name}, column={count_target_col}"
            )